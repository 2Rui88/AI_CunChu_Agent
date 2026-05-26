"""
AI 智能检索模块 —— describe / search / rebuild

describe: 生成文件 AI 描述 + 向量 → 存入 MySQL + FAISS
  - 图片(png/jpg/...) → Qwen-VL 多模态描述
  - 文本(txt/md/py/...) → 下载读取内容作为描述
  - docx → 解压提取 XML 文本作为描述
  - 其他 → 类型 + 文件名拼接描述
search:   自然语言查询 → embedding → FAISS 检索 → 回查 MySQL 返回结果
rebuild:  从 MySQL 全量重建用户 FAISS 索引
"""
import io
import re
import zipfile
import fitz  # PyMuPDF
import openpyxl
from app.chunker import chunk_text, Chunk
import numpy as np
import httpx
from fastapi import APIRouter
from sqlalchemy import select, text
from app.database import Session
from app.models import FileInfo, UserFileList, FileAiDesc, UserFileAiDesc
from app.dependencies import check_token
from app.config import settings
from app.dashscope_client import describe_image, get_embedding, create_client
from app.faiss_service import (
    search as faiss_search,
    add_vector,
    rebuild_from_db,
    l2_normalize,
    get_ntotal,
    load_index,
    is_dirty, clear_dirty,
    acquire_lock, release_lock,
)

router = APIRouter(prefix="/api", tags=["ai"])

# 兼容旧 C 项目的分数阈值
SCORE_THRESHOLD = 0.45


def _build_public_url(db_url: str) -> str:
    """
    将数据库中的文件 URL 替换为公网可访问的 URL。
    DashScope Qwen-VL 需要能直接下载图片，所以必须用公网地址。
    如果未配置 public_server_ip，则回退到 web_server_ip（仅 Docker 内网可用）。
    """
    ip = settings.public_server_ip or settings.web_server_ip
    port = settings.public_server_port or settings.web_server_port

    path_part = db_url
    for prefix in ("http://", "https://"):
        if path_part.startswith(prefix):
            path_part = path_part[len(prefix):]
            pos = path_part.find("/")
            if pos > 0:
                path_part = path_part[pos:]
            break

    return f"http://{ip}:{port}{path_part}"


# ============================================================
#  describe —— AI 文件描述 + 向量化
# ============================================================

@router.post("/ai/describe")
async def ai_describe(body: dict):
    """
    为指定文件生成 AI 描述和向量，存入 file_ai_desc（全局缓存）和
    user_file_ai_desc（用户记录），并追加到用户 FAISS 索引。
    需要: user, token, md5, filename, type, api_key
    """
    user = body.get("user", "")
    token = body.get("token", "")
    if not await check_token(user, token):
        return {"code": 4}

    api_key = body.get("api_key", "")
    if not api_key:
        return {"code": 1, "msg": "missing api_key"}

    md5_val = body.get("md5", "")
    filename = body.get("filename", "")
    file_type = body.get("type", "")
    force = body.get("force", False)

    async with Session() as db:
        # 校验用户是否拥有该文件
        result = await db.execute(
            select(UserFileList).where(
                UserFileList.user == user, UserFileList.md5 == md5_val
            ).limit(1)
        )
        if result.scalar() is None:
            return {"code": 1, "msg": "file not found or no permission"}

        # 已有完成记录且非强制 → 直接返回（检查 chunk_index=0 即可）
        if not force:
            result = await db.execute(
                select(UserFileAiDesc).where(
                    UserFileAiDesc.user == user,
                    UserFileAiDesc.md5 == md5_val,
                    UserFileAiDesc.chunk_index == 0,
                    UserFileAiDesc.status == 1,
                ).limit(1)
            )
            if result.scalar():
                return {"code": 0, "msg": "already exists"}

        # 全局缓存命中 → 复制所有 chunk 到用户表
        if not force:
            result = await db.execute(
                select(FileAiDesc).where(
                    FileAiDesc.md5 == md5_val, FileAiDesc.status == 1
                ).order_by(FileAiDesc.chunk_index)
            )
            caches = result.scalars().all()
            if caches:
                return await _copy_cache_to_user(db, user, md5_val, caches)

        # 生成描述（返回描述文本 + 结构化 metadata 供分块器使用）
        description, metadata = await _generate_description(db, md5_val, filename, file_type, api_key)
        if not description:
            return {"code": 1, "msg": "describe failed"}

        # 分块 + 逐块向量化 + 存储（传入 metadata 供 PDF/Excel 切分器使用）
        chunks = chunk_text(file_type, description, metadata=metadata)
        vecs: list[np.ndarray] = []

        for ch in chunks:
            try:
                embedding = await get_embedding(api_key, ch.text)
            except RuntimeError as exc:
                return {"code": 1, "msg": f"embedding failed: {exc}"}

            vec = np.array(embedding, dtype=np.float32)
            vecs.append(vec)

            # 写入全局缓存（带 chunk_index + context_label）
            await _upsert_file_ai_desc(db, md5_val, ch.text, vec,
                                       ch.index, ch.context_label)
            # 写入用户记录
            await _upsert_user_ai_desc(db, user, md5_val, ch.text, vec,
                                       ch.index, ch.context_label)

        # FAISS 逐块写入（commit 前，保证一致性）
        try:
            for i, ch in enumerate(chunks):
                faiss_id = add_vector(user, vecs[i])
                await db.execute(
                    text(
                        "UPDATE user_file_ai_desc SET faiss_id = :fid "
                        "WHERE user = :user AND md5 = :md5 AND chunk_index = :cidx"
                    ),
                    {"fid": faiss_id, "user": user, "md5": md5_val, "cidx": ch.index},
                )
            await db.commit()
        except Exception:
            return {"code": 1, "msg": "faiss write failed"}

    return {"code": 0, "msg": "ok"}


async def _generate_description(db, md5_val: str, filename: str, file_type: str,
                                api_key: str) -> tuple[str, dict | None]:
    """
    按文件类型生成中文描述，返回 (描述文本, 结构化metadata)。
    metadata 供 chunker 使用（PDF 页面列表、Excel 工作表行等）。
    """
    ft = file_type.lower()

    # ── 图片：Qwen-VL 多模态描述 ──
    if ft in _IMAGE_TYPES:
        result = await db.execute(
            select(FileInfo).where(FileInfo.md5 == md5_val).limit(1)
        )
        fi = result.scalar()
        if fi and fi.url:
            public_url = _build_public_url(fi.url)
            desc = await describe_image(api_key, public_url)
            return desc, None
        return f"图片文件：{filename}", None

    # ── PDF：返回页面列表 + 格式化描述 ──
    if ft == "pdf":
        result = await db.execute(
            select(FileInfo).where(FileInfo.md5 == md5_val).limit(1)
        )
        fi = result.scalar()
        if fi and fi.url:
            internal_url = _build_internal_url(fi.url)
            desc, meta = await _download_and_extract_pdf(internal_url, filename)
            if desc:
                return desc, meta
        return f"PDF文件：{filename}", None

    # ── Excel：返回 sheet 行数据 + 格式化描述 ──
    if ft in ("xlsx", "xls"):
        result = await db.execute(
            select(FileInfo).where(FileInfo.md5 == md5_val).limit(1)
        )
        fi = result.scalar()
        if fi and fi.url:
            internal_url = _build_internal_url(fi.url)
            desc, meta = await _download_and_extract_xlsx(internal_url, filename)
            if desc:
                return desc, meta
        return f"Excel文件：{filename}", None

    # ── 文本类 / docx：下载文件提取内容 ──
    if ft in _TEXT_TYPES or ft == "docx":
        result = await db.execute(
            select(FileInfo).where(FileInfo.md5 == md5_val).limit(1)
        )
        fi = result.scalar()
        if fi and fi.url:
            internal_url = _build_internal_url(fi.url)
            content = await _download_and_extract(internal_url, ft)
            if content:
                return _format_text_description(filename, content), None

    # ── 其他：类型 + 文件名 ──
    return f"{file_type or '未知'}类型的文件：{filename}", None


# ── 支持的文件类型 ──

_IMAGE_TYPES = {"png", "jpg", "jpeg", "gif", "bmp", "webp", "svg"}

_TEXT_TYPES = {
    "txt", "md", "csv", "json", "xml", "html", "htm",
    "log", "c", "cpp", "h", "py", "js", "css", "java",
}


async def _download_and_extract(file_url: str, file_type: str) -> str | None:
    """
    下载文件并提取文本内容。
    - 普通文本文件 → 直接读取（最多 8192 字节）
    - docx → 解压 ZIP 后解析 word/document.xml 中的 <w:t> 标签
    """
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.get(file_url)
            if resp.status_code != 200:
                return None
            data = resp.content
    except Exception:
        return None

    if file_type == "docx":
        return _parse_docx(data)

    # 普通文本：直接解码（尝试 utf-8，失败则 latin-1）
    return _decode_text(data)


def _parse_docx(data: bytes) -> str | None:
    """
    解压 docx（ZIP 格式），从 word/document.xml 中提取 <w:t> 标签内的文本。
    <w:t> 匹配规则：标签名以 w:t 开头，后面是 > 或空格（排除 <w:tab> 等）。
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            xml_bytes = zf.read("word/document.xml")
    except Exception:
        return None

    xml_text = _decode_text(xml_bytes)
    if not xml_text:
        return None

    # 提取所有 <w:t>...</w:t> 标签中的文本内容
    parts = re.findall(r"<w:t[ >][^>]*>(.*?)</w:t>", xml_text)
    text = "".join(parts).strip()
    return text if text else None


def _decode_text(data: bytes) -> str | None:
    """尝试 UTF-8 解码，失败则用 latin-1 兜底"""
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError:
        try:
            return data.decode("latin-1")
        except Exception:
            return None


def _format_text_description(filename: str, content: str) -> str:
    """格式化文本/文档描述：文件名 + 内容（截断到 3000 字符）"""
    desc_len = min(len(content), 3000)
    return f"文件名：{filename}\n文件内容：{content[:desc_len]}"


# ── PDF 文本提取 ──

async def _download_and_extract_pdf(file_url: str, filename: str) -> tuple[str | None, dict | None]:
    """下载 PDF 二进制并提取文本，返回 (描述, {pages, font_sizes})"""
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.get(file_url)
            if resp.status_code != 200:
                return None, None
            data = resp.content
    except Exception:
        return None, None
    return _extract_pdf_text(data, filename)


def _extract_pdf_text(data: bytes, filename: str) -> tuple[str, dict]:
    """
    从 PDF 二进制数据提取文本。
    返回 (描述文本, {"pages": [page_text, ...], "font_sizes": [size, ...]})。
    """
    meta: dict = {"pages": [], "font_sizes": []}
    if not data or len(data) == 0:
        return f"PDF文件（空文件）：{filename}", meta
    try:
        doc = fitz.open(stream=data, filetype="pdf")
    except Exception:
        return f"PDF文件：{filename}", meta
    if doc.is_encrypted:
        doc.close()
        return f"PDF文件（已加密，无法提取文本）：{filename}", meta
    total = doc.page_count
    if total == 0:
        doc.close()
        return f"PDF文件（无页面）：{filename}", meta

    max_chars = settings.pdf_max_extract_chars
    text_parts: list[str] = []
    scanned = 0
    for i in range(total):
        try:
            page = doc.load_page(i)
        except Exception:
            meta["pages"].append(""); meta["font_sizes"].append(0); continue

        page_text = page.get_text("text").strip() if page else ""
        meta["pages"].append(page_text)
        # 提取字号用于突变检测
        try:
            blocks = page.get_text("dict").get("blocks", [])
            fs = 0
            for b in blocks:
                if b.get("type") == 0:
                    for ln in b.get("lines", []):
                        for sp in ln.get("spans", []):
                            fs = round(sp.get("size", 0)); break
                        if fs: break
                if fs: break
            meta["font_sizes"].append(fs)
        except Exception:
            meta["font_sizes"].append(0)

        if not page_text:
            try:
                if page.get_images(): scanned += 1
            except Exception:
                pass
            continue
        text_parts.append(page_text)
        if sum(len(t) for t in text_parts) > max_chars * 2:
            break
    doc.close()

    if not text_parts:
        if scanned > 0:
            return f"PDF文件（疑似扫描型，{scanned}/{total} 页无文本）：{filename}", meta
        return f"PDF文件（无可提取文本）：{filename}", meta

    raw = "\n".join(text_parts)
    raw = _clean_pdf_text(raw)
    if len(raw) > max_chars:
        raw = raw[:max_chars]
        last_period = raw.rfind("。")
        if last_period > max_chars // 2:
            raw = raw[:last_period + 1]
    note = f"（注：{scanned}/{total} 页为扫描页，已跳过）\n" if scanned > 0 else ""
    return f"文件名：{filename}\n{note}PDF内容：{raw}", meta


def _clean_pdf_text(text: str) -> str:
    """清洗 PDF 提取文本：去控制字符、合并空白、规整换行"""
    text = text.replace("\t", " ")
    text = re.sub(r" {2,}", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = "\n".join(line.strip() for line in text.splitlines())
    return text.strip()


# ── Excel 表格提取 ──

async def _download_and_extract_xlsx(file_url: str, filename: str) -> tuple[str | None, dict | None]:
    """下载 Excel 二进制并提取文本，返回 (描述, {sheets: [{name, rows}]})"""
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.get(file_url)
            if resp.status_code != 200:
                return None, None
            data = resp.content
    except Exception:
        return None, None
    return _extract_xlsx_text(data, filename)


def _extract_xlsx_text(data: bytes, filename: str) -> tuple[str, dict]:
    """
    从 Excel 二进制数据提取文本。
    返回 (描述文本, {"sheets": [{"name": ..., "rows": [[...], ...]}, ...]})。
    """
    meta: dict = {"sheets": []}
    if not data or len(data) == 0:
        return f"Excel文件（空文件）：{filename}", meta
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return f"Excel文件（无法解析）：{filename}", meta

    max_chars = settings.pdf_max_extract_chars
    parts: list[str] = []
    sheet_count = 0
    empty_sheets = 0

    for name in wb.sheetnames:
        ws = wb[name]
        sheet_count += 1
        sheet_parts = [f"【{name}】"]
        sheet_rows: list[list[str]] = []
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            values = [str(v) if v is not None else "" for v in row]
            if not any(v for v in values):
                continue
            sheet_rows.append(values)
            sheet_parts.append("\t".join(values))
            row_count += 1
            total = sum(len(p) for p in parts) + sum(len(p) for p in sheet_parts)
            if total > max_chars * 2:
                break
        meta["sheets"].append({"name": name, "rows": sheet_rows})
        if row_count == 0:
            empty_sheets += 1
        else:
            parts.extend(sheet_parts)
    wb.close()

    if not parts:
        return f"Excel文件（{sheet_count} 个工作表均无数据）：{filename}", meta

    raw = "\n".join(parts)
    raw = _clean_pdf_text(raw)
    if len(raw) > max_chars:
        raw = raw[:max_chars]
        last_line = raw.rfind("\n")
        if last_line > max_chars // 2:
            raw = raw[:last_line]
    note = f"（共 {sheet_count} 个工作表）\n" if sheet_count > 1 else ""
    return f"文件名：{filename}\n{note}表格内容：\n{raw}", meta


def _build_internal_url(db_url: str) -> str:
    """
    构造 Docker 内网可访问的文件下载 URL（用于后端下载文本/docx 文件）。
    直接连 MinIO 获取文件（不走 Nginx 避免 HTTPS 重定向）。
    db_url 格式: /files/{bucket}/{object_name}
    """
    path_part = db_url
    # 去掉 /files/{bucket}/ 前缀，得到 MinIO object name
    if path_part.startswith("/files/"):
        # /files/files/ab/cd/name.txt → ab/cd/name.txt
        parts = path_part.split("/", 3)  # ['', 'files', 'files', 'ab/cd/name.txt']
        if len(parts) >= 4:
            path_part = "/" + parts[3]

    return (
        f"http://{settings.minio_endpoint}/{settings.minio_bucket}{path_part}"
    )


async def _copy_cache_to_user(db, user: str, md5_val: str, caches: list) -> dict:
    """从全局缓存 file_ai_desc 复制所有 chunk 到用户表 user_file_ai_desc"""
    for cache in caches:
        if not cache.embedding:
            continue

        vec = np.frombuffer(cache.embedding, dtype=np.float32).copy()
        cidx = getattr(cache, "chunk_index", 0)
        clabel = getattr(cache, "context_label", "")

        await _upsert_user_ai_desc(db, user, md5_val, cache.description,
                                   vec, cidx, clabel)

        faiss_id = add_vector(user, vec)
        await db.execute(
            text(
                "UPDATE user_file_ai_desc SET faiss_id = :fid "
                "WHERE user = :user AND md5 = :md5 AND chunk_index = :cidx"
            ),
            {"fid": faiss_id, "user": user, "md5": md5_val, "cidx": cidx},
        )

    await db.commit()
    return {"code": 0, "msg": "ok"}


async def _upsert_file_ai_desc(db, md5_val: str, description: str, vec: np.ndarray,
                               chunk_index: int = 0, context_label: str = ""):
    """写入或更新全局 AI 描述缓存（按 md5 + chunk_index 唯一）"""
    result = await db.execute(
        select(FileAiDesc).where(
            FileAiDesc.md5 == md5_val, FileAiDesc.chunk_index == chunk_index
        ).limit(1)
    )
    existing = result.scalar()
    if existing:
        existing.description = description
        existing.embedding = vec.tobytes()
        existing.context_label = context_label
        existing.model = settings.vl_model
        existing.status = 1
    else:
        db.add(FileAiDesc(
            md5=md5_val, chunk_index=chunk_index,
            description=description, embedding=vec.tobytes(),
            context_label=context_label, model=settings.vl_model, status=1,
        ))


async def _upsert_user_ai_desc(db, user: str, md5_val: str, description: str,
                                vec: np.ndarray | None, chunk_index: int = 0,
                                context_label: str = ""):
    """写入或更新用户 AI 描述记录（按 user + md5 + chunk_index 唯一）"""
    result = await db.execute(
        select(UserFileAiDesc).where(
            UserFileAiDesc.user == user,
            UserFileAiDesc.md5 == md5_val,
            UserFileAiDesc.chunk_index == chunk_index,
        ).limit(1)
    )
    existing = result.scalar()
    if existing:
        existing.description = description
        if vec is not None:
            existing.embedding = vec.tobytes()
        existing.context_label = context_label
        existing.status = 1
    else:
        db.add(UserFileAiDesc(
            user=user, md5=md5_val, chunk_index=chunk_index,
            description=description,
            embedding=vec.tobytes() if vec is not None else None,
            context_label=context_label, status=1,
        ))


async def _auto_rebuild_if_dirty(user: str):
    """
    脏标记检测 + 自动重建 FAISS 索引。
    文件删除后，dealfile 端点会创建 .dirty 文件，搜索前此处自动重建。
    使用文件锁防止并发搜索触发多次重建。
    """
    lock_fd = acquire_lock(user)
    if lock_fd < 0:
        return  # 加锁失败，放弃本次重建

    try:
        # 二次确认（加锁后其他人可能已清理）
        if not is_dirty(user):
            return

        async with Session() as db:
            result = await db.execute(
                select(UserFileAiDesc).where(
                    UserFileAiDesc.user == user,
                    UserFileAiDesc.status == 1,
                    UserFileAiDesc.embedding.isnot(None),
                ).order_by(UserFileAiDesc.id)
            )
            rows = result.scalars().all()

            if not rows:
                rebuild_from_db(user, [])
            else:
                vectors = [np.frombuffer(r.embedding, dtype=np.float32) for r in rows]
                rebuild_from_db(user, vectors)
                # 回写 faiss_id
                idx = load_index(user)
                for i, r in enumerate(rows):
                    if i < idx.ntotal:
                        r.faiss_id = i
                await db.commit()

        clear_dirty(user)
    finally:
        release_lock(lock_fd)


# ═══════════════════════════════════════════════════════════
#  查询预处理 —— 特征检测 / 分类 / 改写
# ═══════════════════════════════════════════════════════════

# 精确锚点正则（扩展名、日期、编号、英文词）
_RE_PRECISE = re.compile(
    r"\.\w{2,5}$"              # 扩展名 .pdf .xlsx
    r"|\d{4}[年\-]\d{1,2}"     # 日期 2024-03 / 2024年3
    r"|[Qq]\d"                 # 季度 Q1 Q3
    r"|v\d+\.\d+"              # 版本 v2.0 v3.1
    r"|#\d+"                   # 编号 #1024
    r"|[a-zA-Z]{3,}"           # 英文词
)
# 模糊查询改写 prompt
_REWRITE_PROMPT = (
    "将以下自然语言查询改写为3个同义搜索短语，每行一个，只输出短语不要解释。\n"
    "查询: {query}"
)


def _detect_query_features(query: str) -> dict:
    """
    零成本结构特征检测。
    返回 {"is_precise": bool, "keywords": list[str]}。
    命中扩展名/日期/编号等精确锚点时标记 is_precise=True。
    """
    keywords: list[str] = []
    for m in _RE_PRECISE.finditer(query):
        kw = m.group().strip()
        if kw:
            keywords.append(kw)
    return {
        "is_precise": len(keywords) > 0,
        "keywords": keywords,
    }


async def _classify_query(query: str, api_key: str) -> str:
    """
    LLM 查询分类（约 10 token）。
    返回 "模糊" 或 "精确"。
    """
    try:
        client = create_client(api_key)
        resp = await client.chat.completions.create(
            model=settings.chat_model,
            messages=[{
                "role": "user",
                "content": f"以下查询是模糊描述还是精确查找？只回答'模糊'或'精确'。\n查询: {query}",
            }],
            max_tokens=4,
            temperature=0,
        )
        raw = resp.choices[0].message.content.strip()
        return "精确" if "精确" in raw else "模糊"
    except Exception:
        return "模糊"


async def _rewrite_and_embed(query: str, api_key: str) -> np.ndarray:
    """
    对模糊查询生成 3 个变体，分别 embedding，返回均值向量。
    """
    # 生成变体
    try:
        client = create_client(api_key)
        resp = await client.chat.completions.create(
            model=settings.chat_model,
            messages=[{"role": "user", "content": _REWRITE_PROMPT.format(query=query)}],
            max_tokens=100,
            temperature=0.7,
        )
        variants = [v.strip() for v in resp.choices[0].message.content.strip().split("\n") if v.strip()]
        if not variants:
            variants = [query]
    except Exception:
        variants = [query]

    # 分别 embedding 取均值
    vecs = []
    for v in variants[:3]:
        try:
            emb = await get_embedding(api_key, v)
            vecs.append(np.array(emb, dtype=np.float32))
        except Exception:
            pass

    if not vecs:
        emb = await get_embedding(api_key, query)
        return np.array(emb, dtype=np.float32)

    avg = np.mean(vecs, axis=0)
    return l2_normalize(avg)


def _extract_keywords(query: str) -> list[str]:
    """
    从查询中提取精确锚点关键词。
    匹配扩展名(.pdf)、日期(2024-03)、版本号(v2.0)、编号(#1024)、英文词。
    返回去重后的关键词列表。
    """
    seen = set()
    keywords = []
    for m in _RE_PRECISE.finditer(query):
        kw = m.group().strip()
        if kw and kw not in seen:
            seen.add(kw)
            keywords.append(kw)
    return keywords


async def _keyword_search(user: str, keywords: list[str], limit: int = 20) -> list[dict]:
    """
    MySQL 关键词模糊搜索。
    多级排序：命中关键词数量 DESC → 文件名完全匹配优先 → 创建时间 DESC。
    """
    if not keywords:
        return []

    # 构建 WHERE user=? AND (file_name LIKE '%kw1%' OR ...)
    like_parts = " OR ".join([f"ufl.file_name LIKE '%{kw}%'" for kw in keywords])
    sql = (
        f"SELECT ufl.md5, ufl.file_name, fi.url, fi.size, fi.type, ufl.pv, "
        f"  ({' + '.join([f'(ufl.file_name LIKE \"%{kw}%\")' for kw in keywords])}) AS hit_count "
        f"FROM user_file_list ufl "
        f"JOIN file_info fi ON fi.md5 = ufl.md5 "
        f"WHERE ufl.user = :user AND ({like_parts}) "
        f"ORDER BY hit_count DESC, ufl.create_time DESC "
        f"LIMIT :limit"
    )
    try:
        async with Session() as db:
            result = await db.execute(text(sql), {"user": user, "limit": limit})
            rows = result.fetchall()
            files = []
            for row in rows:
                files.append({
                    "md5": row[0], "filename": row[1], "url": row[2],
                    "size": str(row[3] or 0), "type": row[4], "pv": row[5],
                    "hit_count": row[6], "source": "keyword",
                })
            return files
    except Exception:
        return []


# ============================================================
#  search —— AI 语义搜索（含双路召回 + 去重合并）
# ============================================================

@router.post("/ai/search")
async def ai_search(body: dict):
    """
    AI 语义搜索 —— 将查询文本向量化后在用户 FAISS 索引中检索，
    按余弦相似度过滤后回查 MySQL 联表返回结果。
    需要: user, token, query, api_key
    """
    user = body.get("user", "")
    token = body.get("token", "")
    if not await check_token(user, token):
        return {"code": 4}

    query = body.get("query", "")
    api_key = body.get("api_key", "")
    top_k = body.get("top_k", 10)

    if not query or not api_key:
        return {"code": 1, "msg": "missing query or api_key"}

    # 脏标记检测 → 自动重建索引（文件删除后触发）
    if is_dirty(user):
        try:
            await _auto_rebuild_if_dirty(user)
        except Exception:
            pass  # 重建失败不影响搜索，用现存索引继续

    if get_ntotal(user) == 0:
        return {"code": 0, "count": 0, "files": []}

    # ── 查询预处理：特征检测 → 分类 → 改写 ──
    features = _detect_query_features(query)

    # 结构特征未命中时用 LLM 二次分类
    if not features["is_precise"]:
        qtype = await _classify_query(query, api_key)
    else:
        qtype = "精确"

    # 模糊查询 → 改写 + 均值向量；精确查询 → 原始 embedding
    if qtype == "模糊":
        query_vec = await _rewrite_and_embed(query, api_key)
    else:
        try:
            embedding = await get_embedding(api_key, query)
        except RuntimeError as exc:
            return {"code": 1, "msg": f"embedding failed: {exc}"}
        query_vec = np.array(embedding, dtype=np.float32)
    try:
        results = faiss_search(user, query_vec, top_k)
    except Exception as exc:
        return {"code": 1, "msg": f"faiss search failed: {exc}"}

    # ── 关键词路（门控：无精确词则跳过）──
    keywords = _extract_keywords(query)
    kw_files = await _keyword_search(user, keywords, limit=top_k) if keywords else []

    # ── faiss_id → MySQL 联表查询（向量路）──
    async with Session() as db:
        files = []
        seen_md5: set[str] = set()

        for faiss_id, score in results:
            if score < SCORE_THRESHOLD:
                continue
            result = await db.execute(
                select(UserFileAiDesc, UserFileList, FileInfo)
                .join(UserFileList,
                      (UserFileList.user == UserFileAiDesc.user) &
                      (UserFileList.md5 == UserFileAiDesc.md5))
                .join(FileInfo, FileInfo.md5 == UserFileAiDesc.md5)
                .where(
                    UserFileAiDesc.user == user,
                    UserFileAiDesc.faiss_id == faiss_id,
                    UserFileAiDesc.status == 1,
                ).limit(1)
            )
            row = result.first()
            if row:
                uad, ufl, fi = row
                files.append({
                    "md5": uad.md5,
                    "filename": ufl.file_name,
                    "description": uad.description,
                    "url": fi.url,
                    "size": str(fi.size),
                    "type": fi.type,
                    "score": round(score, 4),
                    "source": "向量路",
                })
                seen_md5.add(uad.md5)

        # ── 双路合并：关键词路结果去重追加 ──
        for kf in kw_files:
            if kf["md5"] not in seen_md5:
                kf["score"] = 0.0  # 关键词路无向量分
                files.append(kf)
                seen_md5.add(kf["md5"])

        return {"code": 0, "count": len(files), "files": files}


# ============================================================
#  rebuild —— 重建 FAISS 索引
# ============================================================

@router.post("/ai/rebuild")
async def ai_rebuild(body: dict):
    """
    从 MySQL 中读取用户所有已完成的 AI 描述向量，全量重建 FAISS 索引。
    需要: user, token
    """
    user = body.get("user", "")
    token = body.get("token", "")
    if not await check_token(user, token):
        return {"code": 4}

    async with Session() as db:
        result = await db.execute(
            select(UserFileAiDesc).where(
                UserFileAiDesc.user == user,
                UserFileAiDesc.status == 1,
                UserFileAiDesc.embedding.isnot(None),
            ).order_by(UserFileAiDesc.id)
        )
        rows = result.scalars().all()

        if not rows:
            return {"code": 0, "msg": "rebuilt", "count": 0}

        vectors = [np.frombuffer(r.embedding, dtype=np.float32) for r in rows]
        try:
            rebuild_from_db(user, vectors)
        except Exception as exc:
            return {"code": 1, "msg": f"faiss rebuild failed: {exc}"}

        # 更新 faiss_id（重建后 id 从 0 重新分配）
        idx = load_index(user)
        for i, r in enumerate(rows):
            if i < idx.ntotal:
                r.faiss_id = i
        await db.commit()

    return {"code": 0, "msg": "rebuilt", "count": len(vectors)}
